import os
from typing import Any, List, Tuple
import ast
import json

try:
    import folder_paths
except Exception:  # pragma: no cover
    folder_paths = None


class SaveAttributesToExcel:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "length": ("FLOAT", {"default": 0.0, "step": 0.01}),
                "width": ("FLOAT", {"default": 0.0, "step": 0.01}),
                "height": ("FLOAT", {"default": 0.0, "step": 0.01}),
                "material": ("STRING", {"default": ""}),
                "label": ("STRING", {"default": ""}),
            },
            "optional": {
                "output_dir": (
                    "STRING",
                    {
                        "default": "",
                        "multiline": False,
                        "placeholder": "留空时保存到 ComfyUI outputs 目录",
                    },
                ),
                "filename_prefix": ("STRING", {"default": "attributes", "multiline": False}),
            },
        }

    RETURN_TYPES = ("STRING", "STRING", "STRING")
    RETURN_NAMES = ("xlsx_path", "save_hint", "write_status")
    FUNCTION = "save"
    CATEGORY = "utils/export"

    @staticmethod
    def _parse_list_text(text: str) -> List[str]:
        if not text:
            return []

        raw = text.strip()
        if not raw:
            return []

        if raw.startswith("[") and raw.endswith("]"):
            for parser in (json.loads, ast.literal_eval):
                try:
                    parsed = parser(raw)
                    if isinstance(parsed, (list, tuple)):
                        return [str(item).strip() for item in parsed if str(item).strip()]
                except Exception:
                    continue

        if "\n" in raw:
            return [part.strip() for part in raw.splitlines() if part.strip()]

        return [raw]

    @classmethod
    def _normalize_input_list(cls, value: Any) -> List[str]:
        if value is None:
            return [""]

        if isinstance(value, (list, tuple)):
            items = [str(item).strip() for item in value if str(item).strip()]
            return items if items else [""]

        parsed = cls._parse_list_text(str(value))
        return parsed if parsed else [""]

    @staticmethod
    def _expand_to_length(values: List[str], target_len: int) -> List[str]:
        if target_len <= 0:
            return []
        if not values:
            return [""] * target_len
        if len(values) >= target_len:
            return values[:target_len]
        return values + [values[-1]] * (target_len - len(values))

    @staticmethod
    def _build_write_status(start_row: int, end_row: int, rows_written: int) -> str:
        return (
            f"写入状态: 成功追加 {rows_written} 行，写入行号范围 {start_row}-{end_row}。"
            if rows_written > 0
            else "写入状态: 未写入任何数据。"
        )


    @staticmethod
    def _resolve_output_dir(output_dir: str) -> str:
        custom_dir = (output_dir or "").strip()
        if custom_dir:
            return os.path.abspath(custom_dir)

        if folder_paths is not None and hasattr(folder_paths, "get_output_directory"):
            try:
                return os.path.abspath(folder_paths.get_output_directory())
            except Exception:
                pass

        return os.path.abspath("outputs")

    @staticmethod
    def _resolve_output_path(output_dir: str, filename_prefix: str) -> str:
        safe_output_dir = SaveAttributesToExcel._resolve_output_dir(output_dir)
        os.makedirs(safe_output_dir, exist_ok=True)

        safe_prefix = (filename_prefix or "attributes").strip() or "attributes"
        return os.path.join(safe_output_dir, f"{safe_prefix}.xlsx")

    @staticmethod
    def _build_save_hint(xlsx_path: str) -> str:
        output_dir = os.path.dirname(xlsx_path)
        return f"Excel 已保存到: {xlsx_path}（目录: {output_dir}）"

    def save(
        self,
        length: float,
        width: float,
        height: float,
        material: str,
        label: str,
        output_dir: str = "",
        filename_prefix: str = "attributes",
    ) -> Tuple[str, str, str]:
        try:
            from openpyxl import Workbook, load_workbook
        except Exception as exc:
            raise RuntimeError(
                "需要 openpyxl 才能导出 .xlsx 文件，请在 ComfyUI 环境中安装：pip install openpyxl"
            ) from exc

        out_path = self._resolve_output_path(output_dir, filename_prefix)

        if os.path.exists(out_path):
            wb = load_workbook(out_path)
            ws = wb.active
            if ws.max_row < 1:
                ws.append(["label", "material", "length", "width", "height"])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "data"
            ws.append(["label", "material", "length", "width", "height"])

        label_items = self._normalize_input_list(label)
        material_items = self._normalize_input_list(material)
        length_items = self._normalize_input_list(length)
        width_items = self._normalize_input_list(width)
        height_items = self._normalize_input_list(height)

        rows_to_write = max(
            len(label_items),
            len(material_items),
            len(length_items),
            len(width_items),
            len(height_items),
        )

        label_items = self._expand_to_length(label_items, rows_to_write)
        material_items = self._expand_to_length(material_items, rows_to_write)
        length_items = self._expand_to_length(length_items, rows_to_write)
        width_items = self._expand_to_length(width_items, rows_to_write)
        height_items = self._expand_to_length(height_items, rows_to_write)

        print(
            "输入状态: "
            f"label={label_items}, material={material_items}, length={length_items}, "
            f"width={width_items}, height={height_items}, rows_to_write={rows_to_write}"
        )

        start_row = ws.max_row + 1
        for i in range(rows_to_write):
            ws.append(
                [
                    label_items[i],
                    material_items[i],
                    length_items[i],
                    width_items[i],
                    height_items[i],
                ]
            )
        end_row = ws.max_row

        wb.save(out_path)

        save_hint = self._build_save_hint(out_path)
        write_status = self._build_write_status(start_row, end_row, rows_to_write)
        print(save_hint)
        print(write_status)
        return out_path, save_hint, write_status
